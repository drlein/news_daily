#!/usr/bin/env python3
"""
parse_excel_config.py - 从各处室Excel解析生成三个处室的JSON配置

输出:
  config/工业互联网处.json
  config/医药处.json
  config/原材料金属非金属处.json
"""

import json
import os
import re
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_DIR = os.path.join(BASE_DIR, "config")


def parse_excel(filepath):
    """解析Excel，按处室+类别分组返回"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet1"]
    
    # 按处室分组
    departments = {}
    current_dept = None
    current_category = None
    current_info_type = None
    current_remark = None
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        dept, category, info_type, remark, source = [str(v) if v is not None else "" for v in row]
        
        # 跳过空行
        if not source.strip():
            continue
        
        # 继承上级单元格值（合并单元格情况）
        if dept.strip():
            current_dept = dept.strip()
        if category.strip():
            current_category = category.strip()
        if info_type.strip():
            current_info_type = info_type.strip()
        if remark.strip():
            current_remark = remark.strip()
        
        if not current_dept or not current_category:
            continue
        
        if current_dept not in departments:
            departments[current_dept] = []
        
        departments[current_dept].append({
            "category": current_category,
            "info_type": current_info_type or "",
            "remark": current_remark or "",
            "source": source.strip()
        })
    
    return departments


def generate_config(dept_name, items):
    """为单个处室生成JSON配置"""
    news_sites = []
    sentiment_companies = []
    
    # 搜索关键词（按信息类型生成）
    search_keywords = set()
    
    for item in items:
        source = item["source"]
        category = item["category"]
        info_type = item["info_type"]
        remark = item["remark"]
        
        if category == "新闻资讯类":
            # 判断是网站还是公众号
            if source.startswith("公众号"):
                # 标记公众号，需后续处理
                news_sites.append({
                    "type": "wechat",
                    "name": source.replace("公众号（", "").replace("）", "").strip(),
                    "info_type": info_type,
                    "remark": remark
                })
            else:
                # 提取URL
                url_match = re.search(r'(https?://[^\s)）]+)', source)
                if url_match:
                    url = url_match.group(1).rstrip(")）")
                    news_sites.append({
                        "type": "web",
                        "name": source.split("(")[0].split("（")[0].strip(),
                        "url": url,
                        "info_type": info_type,
                        "remark": remark
                    })
            
            # 按信息类型生成搜索关键词
            if info_type:
                for t in re.split(r'[、，,]', info_type):
                    t = t.strip()
                    if t:
                        search_keywords.add(t)
        
        elif category == "舆情信息类":
            sentiment_companies.append({
                "company": source,
                "info_type": item["info_type"],
                "remark": remark
            })
    
    config = {
        "department": dept_name,
        "news_sources": news_sites,
        "sentiment_companies": sentiment_companies,
        "search_keywords": list(search_keywords),
        "time_range_days": 7,
        "exclude_keywords": ["股市行情", "基金净值", "ETF", "个人理财", "广告"],
        "filter_criteria": {
            "news_priorities": [
                "政策出台与分析",
                "行业重要动态、趋势",
                "重点企业重大项目/投资",
                "产业数据、市场分析",
                "技术创新与科研成果"
            ],
            "sentiment_focus": "根据信息类型和备注判断正/反向舆情"
        }
    }
    
    return config


def save_config(dept_name, config):
    """保存处室配置为JSON文件"""
    # 简化文件名
    safe_name = dept_name.replace("、", "_").replace("，", "_").replace(" ", "")
    filepath = os.path.join(CONFIG_DIR, f"{safe_name}.json")
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    print(f"✅ 生成配置: {filepath}")
    return safe_name


def main():
    excel_path = os.path.join(BASE_DIR, "..", "media/inbound/各处室整合版---df158310-987f-403f-b3e5-47d34063c518.xlsx")
    # 也检查临时文件路径
    alt_path = "/root/.openclaw/media/inbound/各处室整合版---df158310-987f-403f-b3e5-47d34063c518.xlsx"
    
    if os.path.exists(excel_path):
        fp = excel_path
    elif os.path.exists(alt_path):
        fp = alt_path
    else:
        # 尝试找到最新上传的Excel
        import glob
        files = glob.glob("/root/.openclaw/media/inbound/*处室*.xlsx")
        if files:
            fp = sorted(files)[-1]
        else:
            print("❌ 找不到Excel文件")
            return
    
    print(f"📄 解析Excel: {fp}")
    departments = parse_excel(fp)
    
    config_map = {}
    for dept_name, items in departments.items():
        print(f"\n{'='*50}")
        print(f"处室: {dept_name}")
        news_count = sum(1 for i in items if i["category"] == "新闻资讯类")
        sentiment_count = sum(1 for i in items if i["category"] == "舆情信息类")
        print(f"  新闻来源: {news_count}个")
        print(f"  舆情公司: {sentiment_count}家")
        
        config = generate_config(dept_name, items)
        safe_name = save_config(dept_name, config)
        config_map[dept_name] = safe_name
    
    print(f"\n{'='*50}")
    print("✅ 配置生成完成！")
    for dept, sn in config_map.items():
        print(f"  {dept} → config/{sn}.json")
    
    # 输出下一步提示
    print(f"\n📋 下一步运行:")
    for dept, sn in config_map.items():
        print(f"  python3 scripts/step1_fetch.py --config {sn} --date 2026-04-28")


if __name__ == "__main__":
    main()
